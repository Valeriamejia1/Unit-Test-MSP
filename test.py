import unittest
from openpyxl import load_workbook

class TestExcel(unittest.TestCase):
    def assertCellEqual(self, original_cell, new_cell, location):
        self.assertEqual(original_cell.value, new_cell.value, f"Cell value at {location} is different")
   
    def test_UNM(self):
        original_file_path = "C:\\Python tools\\Unit-Test-MSP\\Original Files\\10225354 - COH - Santa Barbara Cottage Hospital.xlsx"
        new_file_path = "C:\\Python tools\\Unit-Test-MSP\\New files\\Cottage Health\\10225354 - Santa Barbara Cottage Hospital.xlsx"

        original_wb = load_workbook(filename=original_file_path)
        new_wb = load_workbook(filename=new_file_path)

        original_sheet_names = original_wb.sheetnames
        new_sheet_names = new_wb.sheetnames

        self.assertEqual(original_sheet_names, new_sheet_names, "Number of sheets are not equal")

        differences_found = []
        for sheet_name in original_sheet_names:
            original_sheet = original_wb[sheet_name]
            new_sheet = new_wb[sheet_name]

            original_rows = original_sheet.iter_rows()
            new_rows = new_sheet.iter_rows()

            for (row_idx, original_row), new_row in zip(enumerate(original_rows), new_rows):
                for col_idx, (original_cell, new_cell) in enumerate(zip(original_row, new_row)):
                    location = f"sheet '{sheet_name}', row {row_idx+1}, column {col_idx+1}"
                    try:
                        self.assertCellEqual(original_cell, new_cell, location)
                    except AssertionError as e:
                        differences_found.append(str(e))

        if differences_found:
            self.fail("\n".join(differences_found))

            
if __name__ == "__main__":
    unittest.main()
