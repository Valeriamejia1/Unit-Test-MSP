import unittest
from openpyxl import load_workbook

class TestExcel(unittest.TestCase):
    def assertCellEqual(self, original_cell, new_cell, location):
        self.assertEqual(original_cell.value, new_cell.value, f"Cell value at {location} is different")
        original_file_path = "C:\\Python tools\\Unit-Test-MSP\\Original Files\\CMH Invoice-10182990-Smith, Sherri A-Nursing Residency ORIG.xlsx"
        new_file_path = "C:\\Python tools\\Unit-Test-MSP\\New files\\Central Maine Healthcare\\Invoice-10182990-Smith, Sherri A-Nursing Residency.xlsx"

        original_wb = load_workbook(filename=original_file_path)
        new_wb = load_workbook(filename=new_file_path)

        original_sheet_names = original_wb.sheetnames
        new_sheet_names = new_wb.sheetnames

        self.assertEqual(original_sheet_names, new_sheet_names, "Number of sheets are not equal")

        differences_found = []
        for sheet_name in original_sheet_names:
            original_sheet = original_wb[sheet_name]
            new_sheet = new_wb[sheet_name]

            original_rows = original_sheet.iter_rows()
            new_rows = new_sheet.iter_rows()

            for (row_idx, original_row), new_row in zip(enumerate(original_rows), new_rows):
                for col_idx, (original_cell, new_cell) in enumerate(zip(original_row, new_row)):
                    location = f"sheet '{sheet_name}', row {row_idx+1}, column {col_idx+1}"
                    try:
                        self.assertCellEqual(original_cell, new_cell, location)
                    except AssertionError as e:
                        differences_found.append(str(e))

        if differences_found:
            self.fail("\n".join(differences_found))

    def test_CSMN(self):
        original_file_path = "C:\\Python tools\\Unit-Test-MSP\\Original Files\\10204890 - CSMN - 1101034 - Wil Rob Adult Primary Care LVN ORIG.xlsx"
        new_file_path = "C:\\Python tools\\Unit-Test-MSP\\New files\\Cedars-Sinai Medical Network\\10204890 - CSMN - 1101034 - IM - Wil Rob Adult Primary Care LVN.xlsx"

        original_wb = load_workbook(filename=original_file_path)
        new_wb = load_workbook(filename=new_file_path)

        original_sheet_names = original_wb.sheetnames
        new_sheet_names = new_wb.sheetnames

        self.assertEqual(original_sheet_names, new_sheet_names, "Number of sheets are not equal")

        differences_found = []
        for sheet_name in original_sheet_names:
            original_sheet = original_wb[sheet_name]
            new_sheet = new_wb[sheet_name]

            original_rows = original_sheet.iter_rows()
            new_rows = new_sheet.iter_rows()

            for (row_idx, original_row), new_row in zip(enumerate(original_rows), new_rows):
                for col_idx, (original_cell, new_cell) in enumerate(zip(original_row, new_row)):
                    location = f"sheet '{sheet_name}', row {row_idx+1}, column {col_idx+1}"
                    try:
                        self.assertCellEqual(original_cell, new_cell, location)
                    except AssertionError as e:
                        differences_found.append(str(e))

        if differences_found:
            self.fail("\n".join(differences_found))

    def test_MVHS(self):
        original_file_path = "C:\\Python tools\\Unit-Test-MSP\\Original Files\\MVHS - Invoice-4102271-ORIG.xlsx"
        new_file_path = "C:\\Python tools\\Unit-Test-MSP\\New files\\MVHS - Wynn Hopsital\\Invoice-4102271.xlsx"

        original_wb = load_workbook(filename=original_file_path)
        new_wb = load_workbook(filename=new_file_path)

        original_sheet_names = original_wb.sheetnames
        new_sheet_names = new_wb.sheetnames

        self.assertEqual(original_sheet_names, new_sheet_names, "Number of sheets are not equal")

        differences_found = []
        for sheet_name in original_sheet_names:
            original_sheet = original_wb[sheet_name]
            new_sheet = new_wb[sheet_name]

            original_rows = original_sheet.iter_rows()
            new_rows = new_sheet.iter_rows()

            for (row_idx, original_row), new_row in zip(enumerate(original_rows), new_rows):
                for col_idx, (original_cell, new_cell) in enumerate(zip(original_row, new_row)):
                    location = f"sheet '{sheet_name}', row {row_idx+1}, column {col_idx+1}"
                    try:
                        self.assertCellEqual(original_cell, new_cell, location)
                    except AssertionError as e:
                        differences_found.append(str(e))

        if differences_found:
            self.fail("\n".join(differences_found))

    def test_ANT(self):
        original_file_path = "C:\\Python tools\\Unit-Test-MSP\\Original Files\\ANTHC-23-D-426757 ORIG.xlsx"
        new_file_path = "C:\\Python tools\\Unit-Test-MSP\\New files\\Alaska Native Trival\\ANTHC-23-D-426757.xlsx"

        original_wb = load_workbook(filename=original_file_path)
        new_wb = load_workbook(filename=new_file_path)

        original_sheet_names = original_wb.sheetnames
        new_sheet_names = new_wb.sheetnames

        self.assertEqual(original_sheet_names, new_sheet_names, "Number of sheets are not equal")

        differences_found = []
        for sheet_name in original_sheet_names:
            original_sheet = original_wb[sheet_name]
            new_sheet = new_wb[sheet_name]

            original_rows = original_sheet.iter_rows()
            new_rows = new_sheet.iter_rows()

            for (row_idx, original_row), new_row in zip(enumerate(original_rows), new_rows):
                for col_idx, (original_cell, new_cell) in enumerate(zip(original_row, new_row)):
                    location = f"sheet '{sheet_name}', row {row_idx+1}, column {col_idx+1}"
                    try:
                        self.assertCellEqual(original_cell, new_cell, location)
                    except AssertionError as e:
                        differences_found.append(str(e))

        if differences_found:
            self.fail("\n".join(differences_found))

    def test_EHS(self):
        original_file_path = "C:\\Python tools\\Unit-Test-MSP\\Original Files\\Eliot 10174064 - OR – Surg Tech ORIG.xlsx"
        new_file_path = "C:\\Python tools\\Unit-Test-MSP\\New files\\Elliot Health System\\10174064 - OR – Surg Tech.xlsx"

        original_wb = load_workbook(filename=original_file_path)
        new_wb = load_workbook(filename=new_file_path)

        original_sheet_names = original_wb.sheetnames
        new_sheet_names = new_wb.sheetnames

        self.assertEqual(original_sheet_names, new_sheet_names, "Number of sheets are not equal")

        differences_found = []
        for sheet_name in original_sheet_names:
            original_sheet = original_wb[sheet_name]
            new_sheet = new_wb[sheet_name]

            original_rows = original_sheet.iter_rows()
            new_rows = new_sheet.iter_rows()

            for (row_idx, original_row), new_row in zip(enumerate(original_rows), new_rows):
                for col_idx, (original_cell, new_cell) in enumerate(zip(original_row, new_row)):
                    location = f"sheet '{sheet_name}', row {row_idx+1}, column {col_idx+1}"
                    try:
                        self.assertCellEqual(original_cell, new_cell, location)
                    except AssertionError as e:
                        differences_found.append(str(e))

        if differences_found:
            self.fail("\n".join(differences_found))

    def test_UWH(self):
        original_file_path = "C:\\Python tools\\Unit-Test-MSP\\Original Files\\UWH Invoice 4212785 - HMC ORIG.xlsx"
        new_file_path = "C:\\Python tools\\Unit-Test-MSP\\New files\\Harborview Medical Center\\Invoice-4212785.xlsx"

        original_wb = load_workbook(filename=original_file_path)
        new_wb = load_workbook(filename=new_file_path)

        original_sheet_names = original_wb.sheetnames
        new_sheet_names = new_wb.sheetnames

        self.assertEqual(original_sheet_names, new_sheet_names, "Number of sheets are not equal")

        differences_found = []
        for sheet_name in original_sheet_names:
            original_sheet = original_wb[sheet_name]
            new_sheet = new_wb[sheet_name]

            original_rows = original_sheet.iter_rows()
            new_rows = new_sheet.iter_rows()

            for (row_idx, original_row), new_row in zip(enumerate(original_rows), new_rows):
                for col_idx, (original_cell, new_cell) in enumerate(zip(original_row, new_row)):
                    location = f"sheet '{sheet_name}', row {row_idx+1}, column {col_idx+1}"
                    try:
                        self.assertCellEqual(original_cell, new_cell, location)
                    except AssertionError as e:
                        differences_found.append(str(e))

        if differences_found:
            self.fail("\n".join(differences_found))

    def test_OSHU(self):
        original_file_path = "C:\\Python tools\\Unit-Test-MSP\\Original Files\\OSHU Invoice 10176150 - Hillsboro Medical Center ORIG.xlsx"
        new_file_path = "C:\\Python tools\\Unit-Test-MSP\\New files\\Oregon Health Science University\\Invoice 10176150 - Hillsboro Medical Center.xlsx"

        original_wb = load_workbook(filename=original_file_path)
        new_wb = load_workbook(filename=new_file_path)

        original_sheet_names = original_wb.sheetnames
        new_sheet_names = new_wb.sheetnames

        self.assertEqual(original_sheet_names, new_sheet_names, "Number of sheets are not equal")

        differences_found = []
        for sheet_name in original_sheet_names:
            original_sheet = original_wb[sheet_name]
            new_sheet = new_wb[sheet_name]

            original_rows = original_sheet.iter_rows()
            new_rows = new_sheet.iter_rows()

            for (row_idx, original_row), new_row in zip(enumerate(original_rows), new_rows):
                for col_idx, (original_cell, new_cell) in enumerate(zip(original_row, new_row)):
                    location = f"sheet '{sheet_name}', row {row_idx+1}, column {col_idx+1}"
                    try:
                        self.assertCellEqual(original_cell, new_cell, location)
                    except AssertionError as e:
                        differences_found.append(str(e))

        if differences_found:
            self.fail("\n".join(differences_found))

    def test_SGMC(self):
        original_file_path = "C:\\Python tools\\Unit-Test-MSP\\Original Files\\SGMC Invoice - #4185042 Cath Lab ORIG.xlsx"
        new_file_path = "C:\\Python tools\\Unit-Test-MSP\\New files\\South Georgia Health System\\Invoice - #4185042 Cath Lab.xlsx"

        original_wb = load_workbook(filename=original_file_path)
        new_wb = load_workbook(filename=new_file_path)

        original_sheet_names = original_wb.sheetnames
        new_sheet_names = new_wb.sheetnames

        self.assertEqual(original_sheet_names, new_sheet_names, "Number of sheets are not equal")

        differences_found = []
        for sheet_name in original_sheet_names:
            original_sheet = original_wb[sheet_name]
            new_sheet = new_wb[sheet_name]

            original_rows = original_sheet.iter_rows()
            new_rows = new_sheet.iter_rows()

            for (row_idx, original_row), new_row in zip(enumerate(original_rows), new_rows):
                for col_idx, (original_cell, new_cell) in enumerate(zip(original_row, new_row)):
                    location = f"sheet '{sheet_name}', row {row_idx+1}, column {col_idx+1}"
                    try:
                        self.assertCellEqual(original_cell, new_cell, location)
                    except AssertionError as e:
                        differences_found.append(str(e))

        if differences_found:
            self.fail("\n".join(differences_found))

    def test_PMH(self):
        original_file_path = "C:\\Python tools\\Unit-Test-MSP\\Original Files\\PMH 10181025 - Manchester Memorial Hospital - ER ORIG.xlsx"
        new_file_path = "C:\\Python tools\\Unit-Test-MSP\\New files\\Eastern Connecticut Health Network Float Pool\\10181025 - Manchester Memorial Hospital - ER.xlsx"

        original_wb = load_workbook(filename=original_file_path)
        new_wb = load_workbook(filename=new_file_path)

        original_sheet_names = original_wb.sheetnames
        new_sheet_names = new_wb.sheetnames

        self.assertEqual(original_sheet_names, new_sheet_names, "Number of sheets are not equal")

        differences_found = []
        for sheet_name in original_sheet_names:
            original_sheet = original_wb[sheet_name]
            new_sheet = new_wb[sheet_name]

            original_rows = original_sheet.iter_rows()
            new_rows = new_sheet.iter_rows()

            for (row_idx, original_row), new_row in zip(enumerate(original_rows), new_rows):
                for col_idx, (original_cell, new_cell) in enumerate(zip(original_row, new_row)):
                    location = f"sheet '{sheet_name}', row {row_idx+1}, column {col_idx+1}"
                    try:
                        self.assertCellEqual(original_cell, new_cell, location)
                    except AssertionError as e:
                        differences_found.append(str(e))

        if differences_found:
            self.fail("\n".join(differences_found))

    def test_UWA(self):
        original_file_path = "C:\\Python tools\\Unit-Test-MSP\\Original Files\\UWA Invoice-4185038 ORIG.xlsx"
        new_file_path = "C:\\Python tools\\Unit-Test-MSP\\New files\\University of Washington Medical Center\\Invoice-4185038.xlsx"

        original_wb = load_workbook(filename=original_file_path)
        new_wb = load_workbook(filename=new_file_path)

        original_sheet_names = original_wb.sheetnames
        new_sheet_names = new_wb.sheetnames

        self.assertEqual(original_sheet_names, new_sheet_names, "Number of sheets are not equal")

        differences_found = []
        for sheet_name in original_sheet_names:
            original_sheet = original_wb[sheet_name]
            new_sheet = new_wb[sheet_name]

            original_rows = original_sheet.iter_rows()
            new_rows = new_sheet.iter_rows()

            for (row_idx, original_row), new_row in zip(enumerate(original_rows), new_rows):
                for col_idx, (original_cell, new_cell) in enumerate(zip(original_row, new_row)):
                    location = f"sheet '{sheet_name}', row {row_idx+1}, column {col_idx+1}"
                    try:
                        self.assertCellEqual(original_cell, new_cell, location)
                    except AssertionError as e:
                        differences_found.append(str(e))

        if differences_found:
            self.fail("\n".join(differences_found))

    def test_CMH(self):
        original_file_path = "C:\\Python tools\\Unit-Test-MSP\\Original Files\\CMH Invoice-10182990-Smith, Sherri A-Nursing Residency ORIG.xlsx"
        new_file_path = "C:\\Python tools\\Unit-Test-MSP\\New files\\Central Maine Healthcare\\Invoice-10182990-Smith, Sherri A-Nursing Residency.xlsx"

        original_wb = load_workbook(filename=original_file_path)
        new_wb = load_workbook(filename=new_file_path)

        original_sheet_names = original_wb.sheetnames
        new_sheet_names = new_wb.sheetnames

        self.assertEqual(original_sheet_names, new_sheet_names, "Number of sheets are not equal")

        differences_found = []
        for sheet_name in original_sheet_names:
            original_sheet = original_wb[sheet_name]
            new_sheet = new_wb[sheet_name]

            original_rows = original_sheet.iter_rows()
            new_rows = new_sheet.iter_rows()

            for (row_idx, original_row), new_row in zip(enumerate(original_rows), new_rows):
                for col_idx, (original_cell, new_cell) in enumerate(zip(original_row, new_row)):
                    location = f"sheet '{sheet_name}', row {row_idx+1}, column {col_idx+1}"
                    try:
                        self.assertCellEqual(original_cell, new_cell, location)
                    except AssertionError as e:
                        differences_found.append(str(e))

    def test_BCH(self):
        original_file_path = "C:\\Python tools\\Unit-Test-MSP\\Original Files\\BCH 10201214 Mease Countryside Hospital.xlsx"
        new_file_path = "C:\\Python tools\\Unit-Test-MSP\\New files\\Bay Care\\10201214 Mease Countryside Hospital.xlsx"

        original_wb = load_workbook(filename=original_file_path)
        new_wb = load_workbook(filename=new_file_path)

        original_sheet_names = original_wb.sheetnames
        new_sheet_names = new_wb.sheetnames

        self.assertEqual(original_sheet_names, new_sheet_names, "Number of sheets are not equal")

        differences_found = []
        for sheet_name in original_sheet_names:
            original_sheet = original_wb[sheet_name]
            new_sheet = new_wb[sheet_name]

            original_rows = original_sheet.iter_rows()
            new_rows = new_sheet.iter_rows()

            for (row_idx, original_row), new_row in zip(enumerate(original_rows), new_rows):
                for col_idx, (original_cell, new_cell) in enumerate(zip(original_row, new_row)):
                    location = f"sheet '{sheet_name}', row {row_idx+1}, column {col_idx+1}"
                    try:
                        self.assertCellEqual(original_cell, new_cell, location)
                    except AssertionError as e:
                        differences_found.append(str(e))

    def test_DCH(self):
        original_file_path = "C:\\Python tools\\Unit-Test-MSP\\Original Files\\DCH 10196613 DCH Regional Medical Center.xlsx"
        new_file_path = "C:\\Python tools\\Unit-Test-MSP\\New files\\DCH Health System\\10196613 DCH Regional Medical Center.xlsx"

        original_wb = load_workbook(filename=original_file_path)
        new_wb = load_workbook(filename=new_file_path)

        original_sheet_names = original_wb.sheetnames
        new_sheet_names = new_wb.sheetnames

        self.assertEqual(original_sheet_names, new_sheet_names, "Number of sheets are not equal")

        differences_found = []
        for sheet_name in original_sheet_names:
            original_sheet = original_wb[sheet_name]
            new_sheet = new_wb[sheet_name]

            original_rows = original_sheet.iter_rows()
            new_rows = new_sheet.iter_rows()

            for (row_idx, original_row), new_row in zip(enumerate(original_rows), new_rows):
                for col_idx, (original_cell, new_cell) in enumerate(zip(original_row, new_row)):
                    location = f"sheet '{sheet_name}', row {row_idx+1}, column {col_idx+1}"
                    try:
                        self.assertCellEqual(original_cell, new_cell, location)
                    except AssertionError as e:
                        differences_found.append(str(e))

    def test_CSMC(self):
        original_file_path = "C:\\Python tools\\Unit-Test-MSP\\Original Files\\CSMC 10189919 Cedars-Sinai Medical Center.xlsx"
        new_file_path = "C:\\Python tools\\Unit-Test-MSP\\New files\\Cedars-Sinai Medical Center\\10189919 Cedars-Sinai Medical Center.xlsx"

        original_wb = load_workbook(filename=original_file_path)
        new_wb = load_workbook(filename=new_file_path)

        original_sheet_names = original_wb.sheetnames
        new_sheet_names = new_wb.sheetnames

        self.assertEqual(original_sheet_names, new_sheet_names, "Number of sheets are not equal")

        differences_found = []
        for sheet_name in original_sheet_names:
            original_sheet = original_wb[sheet_name]
            new_sheet = new_wb[sheet_name]

            original_rows = original_sheet.iter_rows()
            new_rows = new_sheet.iter_rows()

            for (row_idx, original_row), new_row in zip(enumerate(original_rows), new_rows):
                for col_idx, (original_cell, new_cell) in enumerate(zip(original_row, new_row)):
                    location = f"sheet '{sheet_name}', row {row_idx+1}, column {col_idx+1}"
                    try:
                        self.assertCellEqual(original_cell, new_cell, location)
                    except AssertionError as e:
                        differences_found.append(str(e))

        if differences_found:
            self.fail("\n".join(differences_found))

    def test_CSMN(self):
        original_file_path = "C:\\Python tools\\Unit-Test-MSP\\Original Files\\10204890 - CSMN - 1101034 - Wil Rob Adult Primary Care LVN ORIG.xlsx"
        new_file_path = "C:\\Python tools\\Unit-Test-MSP\\New files\\Cedars-Sinai Medical Network\\10204890 - CSMN - 1101034 - IM - Wil Rob Adult Primary Care LVN.xlsx"

        original_wb = load_workbook(filename=original_file_path)
        new_wb = load_workbook(filename=new_file_path)

        original_sheet_names = original_wb.sheetnames
        new_sheet_names = new_wb.sheetnames

        self.assertEqual(original_sheet_names, new_sheet_names, "Number of sheets are not equal")

        differences_found = []
        for sheet_name in original_sheet_names:
            original_sheet = original_wb[sheet_name]
            new_sheet = new_wb[sheet_name]

            original_rows = original_sheet.iter_rows()
            new_rows = new_sheet.iter_rows()

            for (row_idx, original_row), new_row in zip(enumerate(original_rows), new_rows):
                for col_idx, (original_cell, new_cell) in enumerate(zip(original_row, new_row)):
                    location = f"sheet '{sheet_name}', row {row_idx+1}, column {col_idx+1}"
                    try:
                        self.assertCellEqual(original_cell, new_cell, location)
                    except AssertionError as e:
                        differences_found.append(str(e))

        if differences_found:
            self.fail("\n".join(differences_found))

    def test_AH(self):
        original_file_path = "C:\\Python tools\\Unit-Test-MSP\\Original Files\\## ##.xlsx"
        new_file_path = "C:\\Python tools\\Unit-Test-MSP\\New files\\Astria Health\\## ##.xlsx"

        original_wb = load_workbook(filename=original_file_path)
        new_wb = load_workbook(filename=new_file_path)

        original_sheet_names = original_wb.sheetnames
        new_sheet_names = new_wb.sheetnames

        self.assertEqual(original_sheet_names, new_sheet_names, "Number of sheets are not equal")

        differences_found = []
        for sheet_name in original_sheet_names:
            original_sheet = original_wb[sheet_name]
            new_sheet = new_wb[sheet_name]

            original_rows = original_sheet.iter_rows()
            new_rows = new_sheet.iter_rows()

            for (row_idx, original_row), new_row in zip(enumerate(original_rows), new_rows):
                for col_idx, (original_cell, new_cell) in enumerate(zip(original_row, new_row)):
                    location = f"sheet '{sheet_name}', row {row_idx+1}, column {col_idx+1}"
                    try:
                        self.assertCellEqual(original_cell, new_cell, location)
                    except AssertionError as e:
                        differences_found.append(str(e))

        if differences_found:
            self.fail("\n".join(differences_found))

    def test_PVHM(self):
        original_file_path = "C:\\Python tools\\Unit-Test-MSP\\Original Files\\## ##.xlsx"
        new_file_path = "C:\\Python tools\\Unit-Test-MSP\\New files\\Pomona Valley Hospital Medical Center\\## ##.xlsx"

        original_wb = load_workbook(filename=original_file_path)
        new_wb = load_workbook(filename=new_file_path)

        original_sheet_names = original_wb.sheetnames
        new_sheet_names = new_wb.sheetnames

        self.assertEqual(original_sheet_names, new_sheet_names, "Number of sheets are not equal")

        differences_found = []
        for sheet_name in original_sheet_names:
            original_sheet = original_wb[sheet_name]
            new_sheet = new_wb[sheet_name]

            original_rows = original_sheet.iter_rows()
            new_rows = new_sheet.iter_rows()

            for (row_idx, original_row), new_row in zip(enumerate(original_rows), new_rows):
                for col_idx, (original_cell, new_cell) in enumerate(zip(original_row, new_row)):
                    location = f"sheet '{sheet_name}', row {row_idx+1}, column {col_idx+1}"
                    try:
                        self.assertCellEqual(original_cell, new_cell, location)
                    except AssertionError as e:
                        differences_found.append(str(e))

        if differences_found:
            self.fail("\n".join(differences_found))

    def test_EHN(self):
        original_file_path = "C:\\Python tools\\Unit-Test-MSP\\Original Files\\Invoice 10215639 - Highland Park Hospital - Emergency Department.xlsx"
        new_file_path = "C:\\Python tools\\Unit-Test-MSP\\New files\\EHN\\10215639 - Highland Park Hospital - Emergency Department.xlsx"

        original_wb = load_workbook(filename=original_file_path)
        new_wb = load_workbook(filename=new_file_path)

        original_sheet_names = original_wb.sheetnames
        new_sheet_names = new_wb.sheetnames

        self.assertEqual(original_sheet_names, new_sheet_names, "Number of sheets are not equal")

        differences_found = []
        for sheet_name in original_sheet_names:
            original_sheet = original_wb[sheet_name]
            new_sheet = new_wb[sheet_name]

            original_rows = original_sheet.iter_rows()
            new_rows = new_sheet.iter_rows()

            for (row_idx, original_row), new_row in zip(enumerate(original_rows), new_rows):
                for col_idx, (original_cell, new_cell) in enumerate(zip(original_row, new_row)):
                    location = f"sheet '{sheet_name}', row {row_idx+1}, column {col_idx+1}"
                    try:
                        self.assertCellEqual(original_cell, new_cell, location)
                    except AssertionError as e:
                        differences_found.append(str(e))

        if differences_found:
            self.fail("\n".join(differences_found))
       
    def test_ANT(self):
        original_file_path = "C:\\Python tools\\Unit-Test-MSP\\Original Files\\ANTHC-23-D-426757.xlsx"
        new_file_path = "C:\\Python tools\\Unit-Test-MSP\\New files\\Alaska Native Trival\\ANTHC-23-D-426757.xlsx"

        original_wb = load_workbook(filename=original_file_path)
        new_wb = load_workbook(filename=new_file_path)

        original_sheet_names = original_wb.sheetnames
        new_sheet_names = new_wb.sheetnames

        self.assertEqual(original_sheet_names, new_sheet_names, "Number of sheets are not equal")

        differences_found = []
        for sheet_name in original_sheet_names:
            original_sheet = original_wb[sheet_name]
            new_sheet = new_wb[sheet_name]

            original_rows = original_sheet.iter_rows()
            new_rows = new_sheet.iter_rows()

            for (row_idx, original_row), new_row in zip(enumerate(original_rows), new_rows):
                for col_idx, (original_cell, new_cell) in enumerate(zip(original_row, new_row)):
                    location = f"sheet '{sheet_name}', row {row_idx+1}, column {col_idx+1}"
                    try:
                        self.assertCellEqual(original_cell, new_cell, location)
                    except AssertionError as e:
                        differences_found.append(str(e))

        if differences_found:
            self.fail("\n".join(differences_found))
      
    def test_PVMH(self):
        original_file_path = "C:\\Python tools\\Unit-Test-MSP\\Original Files\\10208086 La Verne HC Primary Care, 7002.xlsx"
        new_file_path = "C:\\Python tools\\Unit-Test-MSP\\New files\\Pomona Valley Hospital Medical Center\\10208086 La Verne HC Primary Care, 7002.xlsx"

        original_wb = load_workbook(filename=original_file_path)
        new_wb = load_workbook(filename=new_file_path)

        original_sheet_names = original_wb.sheetnames
        new_sheet_names = new_wb.sheetnames

        self.assertEqual(original_sheet_names, new_sheet_names, "Number of sheets are not equal")

        differences_found = []
        for sheet_name in original_sheet_names:
            original_sheet = original_wb[sheet_name]
            new_sheet = new_wb[sheet_name]

            original_rows = original_sheet.iter_rows()
            new_rows = new_sheet.iter_rows()

            for (row_idx, original_row), new_row in zip(enumerate(original_rows), new_rows):
                for col_idx, (original_cell, new_cell) in enumerate(zip(original_row, new_row)):
                    location = f"sheet '{sheet_name}', row {row_idx+1}, column {col_idx+1}"
                    try:
                        self.assertCellEqual(original_cell, new_cell, location)
                    except AssertionError as e:
                        differences_found.append(str(e))

        if differences_found:
            self.fail("\n".join(differences_found))
        
    def test_CSMC(self):
        original_file_path = "C:\\Python tools\\Unit-Test-MSP\\Original Files\\CSMC 10189919 Cedars-Sinai Medical Center.xlsx"
        new_file_path = "C:\\Python tools\\Unit-Test-MSP\\New files\\Cedars-Sinai Medical Center\\10189919 Cedars-Sinai Medical Center.xlsx"

        original_wb = load_workbook(filename=original_file_path)
        new_wb = load_workbook(filename=new_file_path)

        original_sheet_names = original_wb.sheetnames
        new_sheet_names = new_wb.sheetnames

        self.assertEqual(original_sheet_names, new_sheet_names, "Number of sheets are not equal")

        differences_found = []
        for sheet_name in original_sheet_names:
            original_sheet = original_wb[sheet_name]
            new_sheet = new_wb[sheet_name]

            original_rows = original_sheet.iter_rows()
            new_rows = new_sheet.iter_rows()

            for (row_idx, original_row), new_row in zip(enumerate(original_rows), new_rows):
                for col_idx, (original_cell, new_cell) in enumerate(zip(original_row, new_row)):
                    location = f"sheet '{sheet_name}', row {row_idx+1}, column {col_idx+1}"
                    try:
                        self.assertCellEqual(original_cell, new_cell, location)
                    except AssertionError as e:
                        differences_found.append(str(e))

    def test_CHN(self):
        original_file_path = "C:\\Python tools\\Unit-Test-MSP\\Original Files\\CHN - Invoice - # 10223952 L&D CST NIGHTS.xlsx"
        new_file_path = "C:\\Python tools\\Unit-Test-MSP\\New files\\Community Health Network\\Invoice - # 10223952 L&D CST NIGHTS.xlsx"

        original_wb = load_workbook(filename=original_file_path)
        new_wb = load_workbook(filename=new_file_path)

        original_sheet_names = original_wb.sheetnames
        new_sheet_names = new_wb.sheetnames

        self.assertEqual(original_sheet_names, new_sheet_names, "Number of sheets are not equal")

        differences_found = []
        for sheet_name in original_sheet_names:
            original_sheet = original_wb[sheet_name]
            new_sheet = new_wb[sheet_name]

            original_rows = original_sheet.iter_rows()
            new_rows = new_sheet.iter_rows()

            for (row_idx, original_row), new_row in zip(enumerate(original_rows), new_rows):
                for col_idx, (original_cell, new_cell) in enumerate(zip(original_row, new_row)):
                    location = f"sheet '{sheet_name}', row {row_idx+1}, column {col_idx+1}"
                    try:
                        self.assertCellEqual(original_cell, new_cell, location)
                    except AssertionError as e:
                        differences_found.append(str(e))
                        
        if differences_found:
            self.fail("\n".join(differences_found))

    def test_SMH(self):
        original_file_path = "C:\\Python tools\\Unit-Test-MSP\\Original Files\\SMH - 10203328 - Sturdy Memorial Hospital - A - OR  PACU  ENDO.xlsx"
        new_file_path = "C:\\Python tools\\Unit-Test-MSP\\New files\\Sturdy Memorial Hospital\\10203328 - Sturdy Memorial Hospital - A - OR  PACU  ENDO.xlsx"


        original_wb = load_workbook(filename=original_file_path)
        new_wb = load_workbook(filename=new_file_path)

        original_sheet_names = original_wb.sheetnames
        new_sheet_names = new_wb.sheetnames

        self.assertEqual(original_sheet_names, new_sheet_names, "Number of sheets are not equal")

        differences_found = []
        for sheet_name in original_sheet_names:
            original_sheet = original_wb[sheet_name]
            new_sheet = new_wb[sheet_name]

            original_rows = original_sheet.iter_rows()
            new_rows = new_sheet.iter_rows()

            for (row_idx, original_row), new_row in zip(enumerate(original_rows), new_rows):
                for col_idx, (original_cell, new_cell) in enumerate(zip(original_row, new_row)):
                    location = f"sheet '{sheet_name}', row {row_idx+1}, column {col_idx+1}"
                    try:
                        self.assertCellEqual(original_cell, new_cell, location)
                    except AssertionError as e:
                        differences_found.append(str(e))

        if differences_found:
            self.fail("\n".join(differences_found))
                  
    def test_CSMN(self):
        original_file_path = "C:\\Python tools\\Unit-Test-MSP\\Original Files\\10204890 - CSMN - 1101034 - Wil Rob Adult Primary Care LVN.xlsx"
        new_file_path = "C:\\Python tools\\Unit-Test-MSP\\New files\\Cedars-Sinai Medical Network\\10204890 - CSMN - 1101034 - IM - Wil Rob Adult Primary Care LVN.xlsx"

        original_wb = load_workbook(filename=original_file_path)
        new_wb = load_workbook(filename=new_file_path)

        original_sheet_names = original_wb.sheetnames
        new_sheet_names = new_wb.sheetnames

        self.assertEqual(original_sheet_names, new_sheet_names, "Number of sheets are not equal")

        differences_found = []
        for sheet_name in original_sheet_names:
            original_sheet = original_wb[sheet_name]
            new_sheet = new_wb[sheet_name]

            original_rows = original_sheet.iter_rows()
            new_rows = new_sheet.iter_rows()

            for (row_idx, original_row), new_row in zip(enumerate(original_rows), new_rows):
                for col_idx, (original_cell, new_cell) in enumerate(zip(original_row, new_row)):
                    location = f"sheet '{sheet_name}', row {row_idx+1}, column {col_idx+1}"
                    try:
                        self.assertCellEqual(original_cell, new_cell, location)
                    except AssertionError as e:
                        differences_found.append(str(e))

        if differences_found:
            self.fail("\n".join(differences_found))
   
    def test_EHN(self):
        original_file_path = "C:\\Python tools\\Unit-Test-MSP\\Original Files\\10215642 - Evanston Hospital - Cath Lab  EP Lab ORIG (EHN).xlsx"
        new_file_path = "C:\\Python tools\\Unit-Test-MSP\\New files\\Endeavor Health Northshore\\10215642 - Evanston Hospital - Cath Lab  EP Lab.xlsx"

        original_wb = load_workbook(filename=original_file_path)
        new_wb = load_workbook(filename=new_file_path)

        original_sheet_names = original_wb.sheetnames
        new_sheet_names = new_wb.sheetnames

        self.assertEqual(original_sheet_names, new_sheet_names, "Number of sheets are not equal")

        differences_found = []
        for sheet_name in original_sheet_names:
            original_sheet = original_wb[sheet_name]
            new_sheet = new_wb[sheet_name]

            original_rows = original_sheet.iter_rows()
            new_rows = new_sheet.iter_rows()

            for (row_idx, original_row), new_row in zip(enumerate(original_rows), new_rows):
                for col_idx, (original_cell, new_cell) in enumerate(zip(original_row, new_row)):
                    location = f"sheet '{sheet_name}', row {row_idx+1}, column {col_idx+1}"
                    try:
                        self.assertCellEqual(original_cell, new_cell, location)
                    except AssertionError as e:
                        differences_found.append(str(e))

        if differences_found:
            self.fail("\n".join(differences_found))

    def test_UNM(self):
        original_file_path = "C:\\Python tools\\Unit-Test-MSP\\Original Files\\10215642 - Evanston Hospital - Cath Lab  EP Lab ORIG (EHN).xlsx"
        new_file_path = "C:\\Python tools\\Unit-Test-MSP\\New files\\Endeavor Health Northshore\\10215642 - Evanston Hospital - Cath Lab  EP Lab.xlsx"

        original_wb = load_workbook(filename=original_file_path)
        new_wb = load_workbook(filename=new_file_path)

        original_sheet_names = original_wb.sheetnames
        new_sheet_names = new_wb.sheetnames

        self.assertEqual(original_sheet_names, new_sheet_names, "Number of sheets are not equal")

        differences_found = []
        for sheet_name in original_sheet_names:
            original_sheet = original_wb[sheet_name]
            new_sheet = new_wb[sheet_name]

            original_rows = original_sheet.iter_rows()
            new_rows = new_sheet.iter_rows()

            for (row_idx, original_row), new_row in zip(enumerate(original_rows), new_rows):
                for col_idx, (original_cell, new_cell) in enumerate(zip(original_row, new_row)):
                    location = f"sheet '{sheet_name}', row {row_idx+1}, column {col_idx+1}"
                    try:
                        self.assertCellEqual(original_cell, new_cell, location)
                    except AssertionError as e:
                        differences_found.append(str(e))

        if differences_found:
            self.fail("\n".join(differences_found))

    def test_EIH(self):
        original_file_path = "C:\\Python tools\\Unit-Test-MSP\\Original Files\\10180640 University of New Mexico Hospital.xlsx"
        new_file_path = "C:\\Python tools\\Unit-Test-MSP\\New files\\Eisenhower Medical Center\\10181055 - Unit 2 North LDRP,6400.xlsx"
 
        original_wb = load_workbook(filename=original_file_path)
        new_wb = load_workbook(filename=new_file_path)
 
        original_sheet_names = original_wb.sheetnames
        new_sheet_names = new_wb.sheetnames
 
        self.assertEqual(original_sheet_names, new_sheet_names, "Number of sheets are not equal")
 
        differences_found = []
        for sheet_name in original_sheet_names:
            original_sheet = original_wb[sheet_name]
            new_sheet = new_wb[sheet_name]
 
            original_rows = original_sheet.iter_rows()
            new_rows = new_sheet.iter_rows()
 
            for (row_idx, original_row), new_row in zip(enumerate(original_rows), new_rows):
                for col_idx, (original_cell, new_cell) in enumerate(zip(original_row, new_row)):
                    location = f"sheet '{sheet_name}', row {row_idx+1}, column {col_idx+1}"
                    try:
                        self.assertCellEqual(original_cell, new_cell, location)
                    except AssertionError as e:
                        differences_found.append(str(e))
 
        if differences_found:
            self.fail("\n".join(differences_found))

    def test_UNM(self):
        original_file_path = "C:\\Python tools\\Unit-Test-MSP\\Original Files\\10225354 - COH - Santa Barbara Cottage Hospital.xlsx"
        new_file_path = "C:\\Python tools\\Unit-Test-MSP\\New files\\Cottage Health\\10225354 - Santa Barbara Cottage Hospital.xlsx"

        original_wb = load_workbook(filename=original_file_path)
        new_wb = load_workbook(filename=new_file_path)

        original_sheet_names = original_wb.sheetnames
        new_sheet_names = new_wb.sheetnames

        self.assertEqual(original_sheet_names, new_sheet_names, "Number of sheets are not equal")

        differences_found = []
        for sheet_name in original_sheet_names:
            original_sheet = original_wb[sheet_name]
            new_sheet = new_wb[sheet_name]

            original_rows = original_sheet.iter_rows()
            new_rows = new_sheet.iter_rows()

            for (row_idx, original_row), new_row in zip(enumerate(original_rows), new_rows):
                for col_idx, (original_cell, new_cell) in enumerate(zip(original_row, new_row)):
                    location = f"sheet '{sheet_name}', row {row_idx+1}, column {col_idx+1}"
                    try:
                        self.assertCellEqual(original_cell, new_cell, location)
                    except AssertionError as e:
                        differences_found.append(str(e))

        if differences_found:
            self.fail("\n".join(differences_found))
if __name__ == "__main__":
    unittest.main()
